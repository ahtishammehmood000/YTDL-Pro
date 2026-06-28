"""
metadata.py - Metadata persistence layer for YTDL-Pro.

Contains the three classes that together handle all structured data about
completed downloads:

* :class:`MetadataRecord`  – canonical data container (one per download)
* :class:`HistoryManager`  – reads/writes *history.json* atomically
* :class:`MetadataManager` – appends blocks to *metadata.txt*

These classes are intentionally kept free of any UI or download logic so
they can be imported and used by tests or external scripts without pulling
in yt-dlp or Rich.
"""

from __future__ import annotations

import json
import os
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, Any, Optional

from openpyxl import Workbook, load_workbook

from config import FAILED_FILE, HISTORY_FILE, METADATA_FILE

if TYPE_CHECKING:
    from downloader import Logger


# ===========================================================================
# MetadataRecord
# ===========================================================================


class MetadataRecord:
    """
    Canonical data structure for a single download's metadata.

    Constructed from a raw yt-dlp ``info_dict`` via :meth:`from_info`, then
    consumed by both :class:`HistoryManager` (JSON) and
    :class:`MetadataManager` (plain-text).

    Fields
    ------
    file_number : int
        Sequential output number (matches the stem of ``filename``).
    filename : str
        Final MP4 filename, e.g. ``"3.mp4"``.
    title : str
        Video title as reported by the platform.
    channel : str
        Channel / uploader display name.
    channel_id : str | None
        Platform channel identifier (e.g. YouTube channel ID).
    video_id : str | None
        Platform video identifier (e.g. YouTube video ID).
    duration_seconds : int | None
        Duration in whole seconds.
    duration_formatted : str
        Human-readable duration (``"1h 02m 33s"``).
    upload_date : str | None
        Upload date in ``YYYY-MM-DD`` format.
    upload_timestamp : int | None
        Unix timestamp of the upload, if available.
    resolution : str | None
        Video resolution string (e.g. ``"1920x1080"``).
    fps : float | None
        Frames per second.
    video_codec : str | None
        Video codec string (e.g. ``"avc1.640028"``).
    audio_codec : str | None
        Audio codec string (e.g. ``"mp4a.40.2"``).
    filesize_bytes : int | None
        File size in bytes (exact, sourced from the final file on disk when
        available, otherwise from yt-dlp's reported/estimated value).
    filesize_formatted : str
        Human-readable file size (``"45.2 MB"``).
    description : str | None
        Full, untruncated video description text.
    hashtags : list[str]
        Hashtags merged from all available sources, in priority order
        (description, then tags, then title, then a webpage-metadata
        fallback), deduplicated while preserving first-seen order.
    thumbnail_url : str | None
        URL of the best available thumbnail.
    url : str
        Original URL that was passed to yt-dlp.
    downloaded_at : str
        ISO-8601 UTC timestamp of when the download completed.
    download_status : str
        Always ``"success"`` when constructed via :meth:`from_info`.
    """

    # ------------------------------------------------------------------
    # Construction
    # ------------------------------------------------------------------

    def __init__(
        self,
        *,
        file_number: int,
        filename: str,
        title: str,
        channel: str,
        channel_id: Optional[str],
        video_id: Optional[str],
        duration_seconds: Optional[int],
        duration_formatted: str,
        upload_date: Optional[str],
        upload_timestamp: Optional[int],
        resolution: Optional[str],
        fps: Optional[float],
        video_codec: Optional[str],
        audio_codec: Optional[str],
        filesize_bytes: Optional[int],
        filesize_formatted: str,
        description: Optional[str],
        hashtags: list[str],
        thumbnail_url: Optional[str],
        url: str,
        downloaded_at: str,
        download_status: str,
    ) -> None:
        self.file_number = file_number
        self.filename = filename
        self.title = title
        self.channel = channel
        self.channel_id = channel_id
        self.video_id = video_id
        self.duration_seconds = duration_seconds
        self.duration_formatted = duration_formatted
        self.upload_date = upload_date
        self.upload_timestamp = upload_timestamp
        self.resolution = resolution
        self.fps = fps
        self.video_codec = video_codec
        self.audio_codec = audio_codec
        self.filesize_bytes = filesize_bytes
        self.filesize_formatted = filesize_formatted
        self.description = description
        self.hashtags = hashtags
        self.thumbnail_url = thumbnail_url
        self.url = url
        self.downloaded_at = downloaded_at
        self.download_status = download_status

    # ------------------------------------------------------------------
    # Factory
    # ------------------------------------------------------------------

    @classmethod
    def from_info(
        cls,
        info: dict[str, Any],
        *,
        file_number: int,
        filename: str,
        url: str,
        final_path: Optional[Path] = None,
    ) -> "MetadataRecord":
        """
        Build a :class:`MetadataRecord` from a yt-dlp ``info_dict``.

        When ``final_path`` is supplied and the file exists on disk, exact
        values for filesize, resolution, and codec are taken directly from
        the finished file rather than from yt-dlp's pre-merge estimates.

        Parameters
        ----------
        info:
            The dict returned by ``YoutubeDL.extract_info()``.
        file_number:
            Sequential output number for this download.
        filename:
            Final MP4 filename (e.g. ``"3.mp4"``).
        url:
            The original URL submitted to the downloader.
        final_path:
            Absolute path to the finished MP4 file on disk (optional).
            Used to derive the exact file size after merging.
        """
        raw_duration: Optional[int | float] = info.get("duration")
        duration_seconds: Optional[int] = (
            int(raw_duration) if raw_duration is not None else None
        )

        raw_date: Optional[str] = info.get("upload_date")
        upload_date: Optional[str] = cls._parse_upload_date(raw_date)

        # Prefer explicit timestamp; fall back to deriving it from upload_date.
        upload_timestamp: Optional[int] = info.get("timestamp")
        if upload_timestamp is None and upload_date:
            try:
                dt = datetime.strptime(upload_date, "%Y-%m-%d").replace(
                    tzinfo=timezone.utc
                )
                upload_timestamp = int(dt.timestamp())
            except ValueError:
                pass

        # ----------------------------------------------------------------
        # File size: prefer the real on-disk size of the merged file.
        # ----------------------------------------------------------------
        filesize: Optional[int] = None
        if final_path is not None and final_path.exists():
            try:
                filesize = final_path.stat().st_size
            except OSError:
                pass
        if filesize is None:
            filesize = info.get("filesize") or info.get("filesize_approx")

        # ----------------------------------------------------------------
        # Resolution: prefer the explicit field; derive from width×height
        # if absent.  For the merged output the top-level fields reflect
        # the best video stream selected.
        # ----------------------------------------------------------------
        resolution: Optional[str] = info.get("resolution")
        if not resolution:
            w, h = info.get("width"), info.get("height")
            if w and h:
                resolution = f"{w}x{h}"

        # ----------------------------------------------------------------
        # Codec strings may be composite (e.g. "vp9+opus"); keep as-is.
        # Normalise the literal string "none" that yt-dlp emits for absent
        # streams.
        # ----------------------------------------------------------------
        video_codec: Optional[str] = info.get("vcodec") or None
        audio_codec: Optional[str] = info.get("acodec") or None
        if video_codec == "none":
            video_codec = None
        if audio_codec == "none":
            audio_codec = None

        fps_raw = info.get("fps")
        fps: Optional[float] = float(fps_raw) if fps_raw is not None else None

        hashtags = cls._extract_hashtags(
            description=info.get("description", ""),
            tags=info.get("tags") or [],
            title=info.get("title"),
            webpage_info=info,
        )

        thumbnail_url: Optional[str] = cls._best_thumbnail(info)

        raw_title: str = info.get("title") or "Unknown Title"
        clean_title: str = cls._clean_title(raw_title)

        return cls(
            file_number=file_number,
            filename=filename,
            title=clean_title,
            channel=info.get("uploader") or info.get("channel") or "Unknown Channel",
            channel_id=info.get("channel_id") or info.get("uploader_id"),
            video_id=info.get("id"),
            duration_seconds=duration_seconds,
            duration_formatted=cls._fmt_duration(duration_seconds),
            upload_date=upload_date,
            upload_timestamp=upload_timestamp,
            resolution=resolution,
            fps=fps,
            video_codec=video_codec,
            audio_codec=audio_codec,
            filesize_bytes=filesize,
            filesize_formatted=cls._fmt_size(filesize),
            description=info.get("description"),  # full, never truncated
            hashtags=hashtags,
            thumbnail_url=thumbnail_url,
            url=info.get("webpage_url") or url,
            downloaded_at=datetime.now(tz=timezone.utc).isoformat(),
            download_status="success",
        )

    # ------------------------------------------------------------------
    # Serialisation
    # ------------------------------------------------------------------

    def to_dict(self) -> dict[str, Any]:
        """Return a JSON-serialisable dict of all fields."""
        return {
            "file_number": self.file_number,
            "filename": self.filename,
            "title": self.title,
            "channel": self.channel,
            "channel_id": self.channel_id,
            "video_id": self.video_id,
            "duration_seconds": self.duration_seconds,
            "duration_formatted": self.duration_formatted,
            "upload_date": self.upload_date,
            "upload_timestamp": self.upload_timestamp,
            "resolution": self.resolution,
            "fps": self.fps,
            "video_codec": self.video_codec,
            "audio_codec": self.audio_codec,
            "filesize_bytes": self.filesize_bytes,
            "filesize_formatted": self.filesize_formatted,
            "description": self.description,
            "hashtags": self.hashtags,
            "thumbnail_url": self.thumbnail_url,
            "url": self.url,
            "downloaded_at": self.downloaded_at,
            "download_status": self.download_status,
        }

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_upload_date(raw: Optional[str]) -> Optional[str]:
        """Convert yt-dlp's ``YYYYMMDD`` string to ``YYYY-MM-DD``."""
        if not raw:
            return None
        try:
            return datetime.strptime(raw, "%Y%m%d").strftime("%Y-%m-%d")
        except ValueError:
            return raw

    @staticmethod
    def _fmt_duration(seconds: Optional[int]) -> str:
        if seconds is None:
            return "N/A"
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        if h:
            return f"{h}h {m:02d}m {s:02d}s"
        return f"{m}m {s:02d}s"

    @staticmethod
    def _fmt_size(size_bytes: Optional[int]) -> str:
        if size_bytes is None:
            return "N/A"
        value: float = float(size_bytes)
        for unit in ("B", "KB", "MB", "GB", "TB"):
            if value < 1024:
                return f"{value:.2f} {unit}"
            value /= 1024
        return f"{value:.2f} PB"

    @staticmethod
    def _clean_title(raw_title: str) -> str:
        """
        Return *raw_title* with all ``#hashtag`` tokens removed.

        Hashtags are stripped via regex, then any resulting run of extra
        whitespace is collapsed and the result is trimmed.  If cleaning
        would leave an empty string (e.g. the title was only "#adiaava"),
        the original, uncleaned title is kept instead so the Title field
        is never blank.
        """
        without_hashtags = re.sub(r"#\w+", "", raw_title)
        collapsed = re.sub(r"\s+", " ", without_hashtags).strip()
        return collapsed if collapsed else raw_title

    @staticmethod
    def _extract_hashtags(
        description: Optional[str],
        tags: list[str],
        title: Optional[str] = None,
        webpage_info: Optional[dict[str, Any]] = None,
    ) -> list[str]:
        """
        Return a deduplicated list of hashtags preserving first-seen order.

        Sources are mined in priority order, and results from every source
        are merged together (not just the first non-empty one):

        1. Inline ``#Tag`` patterns in the video description.
        2. Platform tags list (``info["tags"]``).
        3. Inline ``#Tag`` patterns in the video title.
        4. As a last-resort fallback, any inline ``#Tag`` patterns found in
           other webpage-sourced text fields yt-dlp may return for a given
           extractor (currently ``categories``) — only consulted if nothing
           was found in the first three sources, since most extractors
           don't populate this with hashtag-style data.

        Deduplication is case-insensitive; the first-seen casing is kept.
        Alphabetical sorting is intentionally avoided so the original
        order is preserved.
        """
        seen: set[str] = set()
        result: list[str] = []

        def _add(tag: str) -> None:
            key = tag.lower()
            if key not in seen:
                seen.add(key)
                result.append(tag)

        def _add_inline(text: Optional[str]) -> None:
            if not text:
                return
            for match in re.finditer(r"#(\w+)", text):
                _add(f"#{match.group(1)}")

        # 1. Mine inline hashtags from the description (preserves order).
        _add_inline(description)

        # 2. Include explicit platform tags in their original order.
        for tag in tags:
            if tag:
                normalised = tag.strip()
                if normalised:
                    _add(
                        normalised
                        if normalised.startswith("#")
                        else f"#{normalised}"
                    )

        # 3. Mine inline hashtags from the title.
        _add_inline(title)

        # 4. Last resort: scan webpage-sourced metadata fields yt-dlp
        # returns for some extractors, in case a platform stores hashtags
        # somewhere other than description/tags/title.  Only used when
        # nothing has been found yet, since this data is extractor-specific
        # and not guaranteed to exist.
        if not result and webpage_info:
            categories = webpage_info.get("categories") or []
            for category in categories:
                if isinstance(category, str):
                    _add_inline(category)

        return result

    @staticmethod
    def _best_thumbnail(info: dict[str, Any]) -> Optional[str]:
        """
        Select the highest-resolution thumbnail URL from the info dict.

        yt-dlp provides a ``thumbnails`` list ordered by ascending
        preference; we pick the last entry (highest quality) or fall back
        to the top-level ``thumbnail`` key.
        """
        thumbnails: list[dict[str, Any]] = info.get("thumbnails") or []
        if thumbnails:
            with_url = [t for t in thumbnails if t.get("url")]
            if with_url:
                return with_url[-1]["url"]
        return info.get("thumbnail")


# ===========================================================================
# HistoryManager
# ===========================================================================


class HistoryManager:
    """
    Persists a structured JSON log of every completed download to
    *history.json*.

    Each record is a full :class:`MetadataRecord` serialised via
    :meth:`MetadataRecord.to_dict`, providing rich metadata for all
    downstream consumers (CLI ``history`` command, external scripts, etc.).

    The ``downloaded_urls`` property is used by :class:`~downloader.Downloader`
    to skip URLs that have already been processed in previous runs.

    Writes are performed atomically: the JSON payload is written to a
    sibling temporary file first, then renamed over the target.  This
    guarantees that *history.json* is never left in a partially-written
    (corrupt) state if the process crashes mid-write.
    """

    def __init__(
        self,
        path: Path = HISTORY_FILE,
        logger: Optional["Logger"] = None,
    ) -> None:
        self._path = path
        self._log = logger or _default_logger()
        self._records: list[dict[str, Any]] = self._load()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _load(self) -> list[dict[str, Any]]:
        if not self._path.exists():
            return []
        try:
            with self._path.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, list):
                return data
            self._log.warning("history.json has unexpected format; resetting.")
            return []
        except (json.JSONDecodeError, OSError) as exc:
            self._log.error("Cannot read history file: %s", exc)
            return []

    def _save(self) -> None:
        """
        Write records to disk atomically.

        The payload is serialised to a temporary file in the same directory
        as *history.json*, then ``os.replace()`` renames it into place.
        Because ``os.replace()`` is atomic on POSIX (and near-atomic on
        Windows via the same-volume rename guarantee), the target file is
        always either the old version or the new version — never a partial
        write.
        """
        parent = self._path.parent
        parent.mkdir(parents=True, exist_ok=True)

        try:
            # Use a temp file in the same directory to ensure same-filesystem
            # rename (required for atomic replace on most OS/fs combinations).
            fd, tmp_path_str = tempfile.mkstemp(
                dir=parent, prefix=".history_tmp_", suffix=".json"
            )
            tmp_path = Path(tmp_path_str)
            try:
                with os.fdopen(fd, "w", encoding="utf-8") as fh:
                    json.dump(self._records, fh, indent=2, ensure_ascii=False)
                    fh.flush()
                    os.fsync(fh.fileno())
                os.replace(tmp_path, self._path)
            except Exception:
                # Clean up the orphaned temp file on any error.
                try:
                    tmp_path.unlink(missing_ok=True)
                except OSError:
                    pass
                raise
        except OSError as exc:
            self._log.error("Cannot write history file: %s", exc)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    @property
    def downloaded_urls(self) -> set[str]:
        """Return the set of URLs that have already been downloaded."""
        return {r["url"] for r in self._records}

    def record(self, meta: MetadataRecord) -> None:
        """
        Append the full metadata record for a successful download and
        persist the updated list to disk atomically.

        Parameters
        ----------
        meta:
            Populated :class:`MetadataRecord` instance.
        """
        self._records.append(meta.to_dict())
        self._save()
        self._log.debug(
            "History updated: %s → %s (video_id=%s)",
            meta.url,
            meta.filename,
            meta.video_id,
        )

    def all_records(self) -> list[dict[str, Any]]:
        """Return a shallow copy of all history records."""
        return list(self._records)


# ===========================================================================
# MetadataManager
# ===========================================================================


class MetadataManager:
    """
    Appends a structured, human-readable metadata block to *metadata.txt*
    after each successful download.

    The block uses clearly labelled sections enclosed between separator
    rules, making the file readable in any plain-text viewer while remaining
    trivially parseable by scripts.  Descriptions are stored in full — never
    truncated.
    """

    _OUTER_SEP: str = "=" * 72
    _INNER_SEP: str = "-" * 72

    def __init__(
        self,
        path: Path = METADATA_FILE,
        logger: Optional["Logger"] = None,
        videos_dir: Optional[Path] = None,
    ) -> None:
        """
        Parameters
        ----------
        path:
            Explicit metadata file location.  Used as-is when *videos_dir*
            is not supplied — kept for backward compatibility with any
            existing caller that constructs ``MetadataManager`` directly.
        logger:
            Optional :class:`Logger`; falls back to :func:`_default_logger`.
        videos_dir:
            Directory where this run's videos are being saved.  When
            given, metadata.txt is always written inside this exact
            folder (``videos_dir / "metadata.txt"``), regardless of
            *path*.  This is what makes metadata.txt follow the user's
            chosen download destination instead of always landing in a
            single fixed project-root location.
        """
        self._path = (videos_dir / "metadata.txt") if videos_dir is not None else path
        self._log = logger or _default_logger()

    def append(self, meta: MetadataRecord) -> None:
        """
        Write a structured metadata block for a completed download.

        Parameters
        ----------
        meta:
            Populated :class:`MetadataRecord` instance.
        """
        hashtag_str = (
            "  ".join(meta.hashtags) if meta.hashtags else "No hashtags"
        )
        description_block = self._format_description(meta.description)

        lines: list[str] = [
            self._OUTER_SEP,
            "  YTDL-Pro  |  Download Record",
            self._OUTER_SEP,
            "",
            # ── General Information ─────────────────────────────────────
            "  General Information",
            self._INNER_SEP,
            f"  File Number      : {meta.file_number}",
            f"  Filename         : {meta.filename}",
            f"  Original URL     : {meta.url}",
            "",
            # ── Video Information ────────────────────────────────────────
            "  Video Information",
            self._INNER_SEP,
            f"  Title            : {meta.title}",
            f"  Channel          : {meta.channel}",
            f"  Channel ID       : {meta.channel_id or 'N/A'}",
            f"  Video ID         : {meta.video_id or 'N/A'}",
            f"  Upload Date      : {meta.upload_date or 'N/A'}",
            f"  Upload Timestamp : {meta.upload_timestamp if meta.upload_timestamp is not None else 'N/A'}",
            f"  Thumbnail URL    : {meta.thumbnail_url or 'N/A'}",
            "",
            # ── Technical Information ────────────────────────────────────
            "  Technical Information",
            self._INNER_SEP,
            f"  Duration         : {meta.duration_formatted}",
            f"  Duration (sec)   : {meta.duration_seconds if meta.duration_seconds is not None else 'N/A'}",
            f"  Resolution       : {meta.resolution or 'N/A'}",
            f"  FPS              : {meta.fps if meta.fps is not None else 'N/A'}",
            f"  Video Codec      : {meta.video_codec or 'N/A'}",
            f"  Audio Codec      : {meta.audio_codec or 'N/A'}",
            f"  File Size        : {meta.filesize_formatted}",
            f"  File Size (bytes): {meta.filesize_bytes if meta.filesize_bytes is not None else 'N/A'}",
            "",
            # ── Hashtags ─────────────────────────────────────────────────
            "  Hashtags",
            self._INNER_SEP,
            f"  {hashtag_str}",
            "",
            # ── Description ──────────────────────────────────────────────
            "  Description",
            self._INNER_SEP,
            description_block,
            "",
            # ── Download Information ──────────────────────────────────────
            "  Download Information",
            self._INNER_SEP,
            f"  Downloaded At    : {meta.downloaded_at}",
            f"  Download Status  : {meta.download_status}",
            "",
            self._OUTER_SEP,
            "",
        ]

        try:
            self._path.parent.mkdir(parents=True, exist_ok=True)
            with self._path.open("a", encoding="utf-8") as fh:
                fh.write("\n".join(lines))
        except OSError as exc:
            self._log.error("Cannot write metadata file: %s", exc)

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _format_description(description: Optional[str]) -> str:
        """
        Indent every line of the description for visual separation inside
        the block.  Returns the full text — no truncation is applied.
        """
        if not description:
            return "  (no description)"
        text = description.strip()
        # Indent each line consistently; preserve blank lines as-is.
        return "\n".join(
            f"  {line}" if line.strip() else ""
            for line in text.splitlines()
        )


# ===========================================================================
# ExcelMetadataManager
# ===========================================================================


class ExcelMetadataManager:
    """
    Appends one row to *metadata.xlsx* after each successful download.

    This mirrors :class:`MetadataManager` (which writes *metadata.txt*) but
    targets a spreadsheet instead, so the two files stay synchronized: every
    call to :meth:`append` should be paired with a call to
    ``MetadataManager.append`` for the same record.

    Behaviour:

    * If *metadata.xlsx* does not exist yet, it is created with the header
      row below.
    * If it already exists, the workbook is loaded as-is and a new row is
      appended at the end — existing rows are never modified, reordered,
      or overwritten.
    * Saving is done by rewriting the whole workbook file; this is the
      normal openpyxl save pattern and does not alter previously written
      rows.
    """

    HEADERS: tuple[str, ...] = (
        "File No",
        "File Name",
        "Title",
        "Hashtags",
        "Channel",
        "Upload Date",
        "Duration",
        "Resolution",
        "FPS",
        "Video Codec",
        "Audio Codec",
        "File Size (MB)",
        "URL",
        "Downloaded At",
    )

    _SHEET_NAME: str = "Metadata"

    def __init__(
        self,
        path: Optional[Path] = None,
        logger: Optional["Logger"] = None,
        videos_dir: Optional[Path] = None,
    ) -> None:
        """
        Parameters
        ----------
        path:
            Explicit metadata.xlsx location.  Used as-is when *videos_dir*
            is not supplied.
        logger:
            Optional :class:`Logger`; falls back to :func:`_default_logger`.
        videos_dir:
            Directory where this run's videos are being saved.  When
            given, metadata.xlsx is always written inside this exact
            folder (``videos_dir / "metadata.xlsx"``), mirroring
            :class:`MetadataManager`'s placement of metadata.txt.
        """
        if videos_dir is not None:
            self._path = videos_dir / "metadata.xlsx"
        elif path is not None:
            self._path = path
        else:
            self._path = METADATA_FILE.with_name("metadata.xlsx")
        self._log = logger or _default_logger()

    def append(self, meta: MetadataRecord) -> None:
        """
        Append a single row for a completed download to *metadata.xlsx*.

        Creates the workbook with headers first if it does not yet exist.
        Existing rows are left untouched — the row for this download is
        always added after the last existing row.
        """
        hashtag_str = "  ".join(meta.hashtags) if meta.hashtags else "No hashtags"
        filesize_mb: Optional[float] = (
            round(meta.filesize_bytes / (1024 * 1024), 2)
            if meta.filesize_bytes is not None
            else None
        )

        row = (
            meta.file_number,
            meta.filename,
            meta.title,
            hashtag_str,
            meta.channel,
            meta.upload_date or "N/A",
            meta.duration_formatted,
            meta.resolution or "N/A",
            meta.fps if meta.fps is not None else "N/A",
            meta.video_codec or "N/A",
            meta.audio_codec or "N/A",
            filesize_mb if filesize_mb is not None else "N/A",
            meta.url,
            meta.downloaded_at,
        )

        try:
            self._path.parent.mkdir(parents=True, exist_ok=True)

            if self._path.exists():
                workbook = load_workbook(self._path)
                if self._SHEET_NAME in workbook.sheetnames:
                    sheet = workbook[self._SHEET_NAME]
                else:
                    # File exists but lacks our sheet (unexpected edit by
                    # the user) — add it rather than touching anything else.
                    sheet = workbook.create_sheet(self._SHEET_NAME)
                    sheet.append(self.HEADERS)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = self._SHEET_NAME
                sheet.append(self.HEADERS)

            sheet.append(row)
            workbook.save(self._path)
        except OSError as exc:
            self._log.error("Cannot write Excel metadata file: %s", exc)


# ===========================================================================
# Internal helper
# ===========================================================================


def _default_logger() -> Any:
    """
    Lazy import of Logger to avoid a circular dependency at module load time.

    ``metadata`` imports nothing from ``downloader``; ``downloader`` imports
    from ``metadata``.  The only coupling is Logger, which HistoryManager and
    MetadataManager accept as an *optional* parameter.  When no logger is
    provided they fall back to this function, which imports Logger at call
    time (after all modules are fully loaded).
    """
    from downloader import Logger  # noqa: PLC0415  (intentional lazy import)

    return Logger()
